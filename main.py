import os
import sys
from PySide6 import *
from PySide6.QtCore import QFileInfo
from PySide6.QtGui import Qt, QIcon, QPixmap, QFont, QBrush
from PySide6.QtUiTools import loadUiType
from PySide6.QtWidgets import (
    QMainWindow,
    QDialog,
    QWidget,
    QFileDialog,
    QTableWidgetItem,
    QApplication,
)
import openpyxl
import pymysql
from pymysql.constants import CLIENT
from db.db import registerWarehouse
from result.result_tab import ResultTab
from simulation.simulation_tab import SimulationTab
from simulator.pathfinding import registerMap

conn = pymysql.connect(
    host="127.0.0.1",
    port=3306,
    user="root",
    password="1290",
    db="lghpdb",
    charset="utf8",
    client_flag=CLIENT.MULTI_STATEMENTS,
    autocommit=True,
)
cur = conn.cursor()

current_path = os.path.dirname(os.path.abspath(__file__))
r_path = os.path.join(current_path, "image", "r_arrow.png")
u_path = os.path.join(current_path, "image", "u_arrow.png")
l_path = os.path.join(current_path, "image", "l_arrow.png")
d_path = os.path.join(current_path, "image", "d_arrow.png")
r_l_path = os.path.join(current_path, "image", "r_l_arrow.png")
u_d_path = os.path.join(current_path, "image", "u_d_arrow.png")
a_path = os.path.join(current_path, "image", "all_arrow.png")
logo_path=os.path.join(current_path, "image","logo_trans.png")

def resource_path(relative_path):
    base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

def dir_img(path, text):
    pixmap = QPixmap(path)
    pixmap.scaled(1, 1)
    icon = QIcon(pixmap)
    item = QTableWidgetItem()
    item.setText(text)
    item.setIcon(icon)
    font = QFont()
    font.setPointSize(1)
    item.setFont(font)
    item.setForeground(QBrush(Qt.transparent))
    item.setTextAlignment(Qt.AlignCenter)
    return item

form = resource_path("homePage.ui")
form_class = loadUiType(form)[0]
form_second = resource_path("simul.ui")
form_secondwindow = loadUiType(form_second)[0]


class WindowClass(QMainWindow, form_class):
    def __init__(self):
        global projectid
        projectid = "tmp"
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("시뮬레이터")
        self.setWindowIcon(QIcon('./image/logo.png'))
        self.setGeometry(100, 50, 1000, 550)
        pixmap = QPixmap(logo_path)
        self.logolabel.setPixmap(pixmap)
        self.loadMap.clicked.connect(self.btn_loadMap)  # loadMap button 클릭

    # -loadMap button 함수: simul.ui로 창전환
    def btn_loadMap(self):
        global projectid
        cur.execute("CALL deleteProject(%s)", [projectid])
        self.hide()
        self.second = secondwindow()
        self.second.exec()
        self.show()


class secondwindow(QDialog, QWidget, form_secondwindow):
    def __init__(self):
        super(secondwindow, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("시뮬레이터")
        self.setWindowIcon(QIcon('./image/logo.png'))
        self.setGeometry(100, 50, 1000, 550)
        self.show()

        ###   overview  tab ###
        # overview-1.map 파일 미리보기
        file = QFileDialog.getOpenFileName(self, "", "", "xlsx파일 (*.xlsx);; All File(*)")  # !!저장파일 타입 정해지면, 확장자에 추가
        global filename  # 선언, 할당 분리
        filename = file[0]
        load_xlsx = openpyxl.load_workbook(file[0], data_only=True)
        load_sheet = load_xlsx["NewSheet1"]

        # 파일 이름으로 db에서 해당 정보 연결
        global file_name, simul_name, s_count
        file_name = QFileInfo(file[0]).baseName()
        # 우선 프로젝트명 : p1, 시뮬명 : s1으로 생성, 이후에 값들 입력받으면 변경
        s_count = int(1)
        simul_name = str(file_name) + "_s" + str(s_count)
        s_count = s_count + 1
        sql = "CALL deleteProject('p1'); CALL createProject('p1', NULL, NULL, NULL); CALL createSimul('p1', %s); CALL updateGridtoSimul(%s, %s); CALL updateProjectRunning(1, 'p1');"
        cur.execute(sql, [simul_name, simul_name, str(file_name)])
        # db에서 가져온 GridSize로 row, col 변경
        sql = "SELECT GridSizeX FROM grid " + "WHERE Grid_ID = %s;"
        cur.execute(sql, [str(file_name)])
        file_col = cur.fetchone()
        sql = "SELECT GridSizeY FROM grid " + "WHERE Grid_ID = %s;"
        cur.execute(sql, [str(file_name)])
        file_row = cur.fetchone()
        row = int(file_row[0])
        col = int(file_col[0])

        # park
        registerWarehouse(file_name)
        registerMap()
        self.tabWidget.addTab(SimulationTab(), "Simulations")
        self.tabWidget.addTab(ResultTab(), "Results")
        # park

        self.map.setColumnCount(col)
        self.map.setRowCount(row)
        for i in range(row):
            for j in range(col):
                self.map.setItem(i, j, QTableWidgetItem())
        self.map.resizeColumnsToContents()
        self.map.resizeRowsToContents()

        self.map.setStyleSheet("background-color:white")
        self.setStyleSheet("background-color:rgb(1,35,38);")
        cnum = 1
        for i in range(1, row + 1):
            for j in range(1, col + 1):
                cell_num = str(file_name) + '_c' + str(cnum).zfill(4)
                cnum = cnum + 1
                sql = "SELECT * FROM cell " + "WHERE Cell_ID = %s;"
                cur.execute(sql, [cell_num])
                cellinfo = cur.fetchone()
                if cellinfo[6] == 1:
                    if cellinfo[7] == 1:
                        item = dir_img(u_d_path, "↕")
                        self.map.setItem(i - 1, j - 1, item)
                        if cellinfo[8] == 1:
                            item = dir_img(a_path, "↕↔")
                            self.map.setItem(i - 1, j - 1, item)
                    else:
                        item = dir_img(u_path, "↑")
                        self.map.setItem(i - 1, j - 1, item)
                elif cellinfo[7] == 1:
                    item = dir_img(d_path, "↓")
                    self.map.setItem(i - 1, j - 1, item)
                elif cellinfo[8] == 1:
                    if cellinfo[9] == 1:
                        item = dir_img(r_l_path, "↔")
                        self.map.setItem(i - 1, j - 1, item)
                    else:
                        item = dir_img(l_path, "←")
                        self.map.setItem(i - 1, j - 1, item)
                elif cellinfo[9] == 1:
                    item = dir_img(r_path, "→")
                    self.map.setItem(i - 1, j - 1, item)

                if load_sheet.cell(i, j).fill.start_color.index == 'FFFFFF00':
                    self.map.item(i - 1, j - 1).setBackground(Qt.yellow)
                    self.map.item(i - 1, j - 1).setForeground(Qt.black)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FF0000FF':
                    self.map.item(i - 1, j - 1).setBackground(Qt.darkBlue)
                    self.map.item(i - 1, j - 1).setForeground(Qt.white)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FF008000':
                    self.map.item(i - 1, j - 1).setBackground(Qt.darkGreen)
                    self.map.item(i - 1, j - 1).setForeground(Qt.white)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FFFF0000':
                    self.map.item(i - 1, j - 1).setBackground(Qt.red)
                    self.map.item(i - 1, j - 1).setForeground(Qt.white)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FF808080':
                    self.map.item(i - 1, j - 1).setBackground(Qt.darkGray)
                    self.map.item(i - 1, j - 1).setForeground(Qt.white)
                else:
                    self.map.item(i - 1, j - 1).setForeground(Qt.black)

        # overview-2.속성별 색상 정보 보여주기
        sql = "SELECT * FROM grid " + "WHERE Grid_ID = %s;"
        cur.execute(sql, [str(file_name)])
        file_grid = cur.fetchone()

        # 충전셀 색상
        if file_grid[13] == 1:
            self.color_charge.setText("Yellow")
            self.c_charge.setStyleSheet("background:yellow")
        elif file_grid[13] == 2:
            self.color_charge.setText("Red")
            self.c_charge.setStyleSheet("background:red")
        elif file_grid[13] == 3:
            self.color_charge.setText("Green")
            self.c_charge.setStyleSheet("background:green")
        elif file_grid[13] == 4:
            self.color_charge.setText("Blue")
            self.c_charge.setStyleSheet("background:blue")
        elif file_grid[13] == 5:
            self.color_charge.setText("Grey")
            self.c_charge.setStyleSheet("background:darkgrey")

        # 슈트셀 색상
        if file_grid[14] == 1:
            self.color_chute.setText("Yellow")
            self.c_chute.setStyleSheet("background:yellow")
        elif file_grid[14] == 2:
            self.color_chute.setText("Red")
            self.c_chute.setStyleSheet("background:red")
        elif file_grid[14] == 3:
            self.color_chute.setText("Green")
            self.c_chute.setStyleSheet("background:green")
        elif file_grid[14] == 4:
            self.color_chute.setText("Blue")
            self.c_chute.setStyleSheet("background:blue")
        elif file_grid[14] == 5:
            self.color_chute.setText("Grey")
            self.c_chute.setStyleSheet("background:darkgrey")

        # 워크스테이션셀 색상
        if file_grid[15] == 1:
            self.color_ws.setText("Yellow")
            self.c_ws.setStyleSheet("background:yellow")
        elif file_grid[15] == 2:
            self.color_ws.setText("Red")
            self.c_ws.setStyleSheet("background:red")
        elif file_grid[15] == 3:
            self.color_ws.setText("Green")
            self.c_ws.setStyleSheet("background:green")
        elif file_grid[15] == 4:
            self.color_ws.setText("Blue")
            self.c_ws.setStyleSheet("background:blue")
        elif file_grid[15] == 5:
            self.color_ws.setText("Grey")
            self.c_ws.setStyleSheet("background:darkgrey")

        # 버퍼셀 색상
        if file_grid[16] == 1:
            self.color_buffer.setText("Yellow")
            self.c_buffer.setStyleSheet("background:yellow")
        elif file_grid[16] == 2:
            self.color_buffer.setText("Red")
            self.c_buffer.setStyleSheet("background:red")
        elif file_grid[16] == 3:
            self.color_buffer.setText("Green")
            self.c_buffer.setStyleSheet("background:green")
        elif file_grid[16] == 4:
            self.color_buffer.setText("Blue")
            self.c_buffer.setStyleSheet("background:blue")
        elif file_grid[16] == 5:
            self.color_buffer.setText("Grey")
            self.c_buffer.setStyleSheet("background:darkgrey")

        # 블락셀 색상
        if file_grid[17] == 1:
            self.color_block.setText("Yellow")
            self.c_block.setStyleSheet("background:yellow")
        elif file_grid[17] == 2:
            self.color_block.setText("Red")
            self.c_block.setStyleSheet("background:red")
        elif file_grid[17] == 3:
            self.color_block.setText("Green")
            self.c_block.setStyleSheet("background:green")
        elif file_grid[17] == 4:
            self.color_block.setText("Blue")
            self.c_block.setStyleSheet("background:blue")
        elif file_grid[17] == 5:
            self.color_block.setText("Grey")
            self.c_block.setStyleSheet("background:darkgrey")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()

    app.exec()
cur.execute("CALL deleteProject(%s)", [projectid])

conn.close()